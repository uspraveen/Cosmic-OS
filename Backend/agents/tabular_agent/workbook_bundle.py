"""Deterministic workbook -> parquet + DuckDB + manifest pipeline (tabular specialist)."""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd

from shared.tabular_artifacts import validate_safe_sheet_id

from .config import TabularAgentConfig


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def sanitize_sheet_id(name: str, index: int) -> str:
    raw = re.sub(r"[^a-zA-Z0-9_]+", "_", (name or f"sheet_{index}").strip())[:80].strip("_").lower()
    return raw or f"sheet_{index}"


def csv_sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        return dialect.delimiter
    except Exception:
        return ","


def _sniff_csv_file(path: Path) -> tuple[str, str]:
    from charset_normalizer import from_bytes

    raw = path.read_bytes()
    best = from_bytes(raw).best()
    encoding = best.encoding if best is not None else "utf-8"
    try:
        text = raw.decode(encoding)
    except UnicodeDecodeError:
        encoding = "utf-8"
        text = raw.decode(encoding, errors="replace")
    lines = text.splitlines()
    sample = "\n".join(lines[: min(40, len(lines))])
    sep = csv_sniff_delimiter(sample)
    return encoding, sep


def _trim_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, int, int]:
    if df.empty:
        return df, 0, 0
    mask = df.notna() & (df.astype(str).str.strip() != "")
    row_any = mask.any(axis=1)
    if not row_any.any():
        return df, 0, 0
    first_row = int(row_any.idxmax())
    last_row = int(row_any[::-1].idxmax())
    col_any = mask.any(axis=0)
    cols_idx = [i for i, v in enumerate(col_any.values) if v]
    if not cols_idx:
        return df.iloc[first_row : last_row + 1], first_row, 0
    first_col, last_col = cols_idx[0], cols_idx[-1]
    trimmed = df.iloc[first_row : last_row + 1, first_col : last_col + 1].copy()
    return trimmed, first_row, first_col


HEADER_V1_NOTES = (
    "Heuristic v1: after trimming empty margins, the first row of the remaining block is treated as the "
    "column header; multi-row headers, title rows above the table, and multiple tables on one sheet are not "
    "detected in this build."
)


def build_header_detection_v1(
    *,
    header_idx: int,
    data_start: int,
) -> dict[str, Any]:
    return {
        "detection_mode": "v1_first_row_after_trim",
        "detected_header_row": header_idx,
        "detected_data_start_row": data_start,
        "multi_row_header_applied": False,
        "multiple_table_warning": False,
        "notes": [HEADER_V1_NOTES],
    }


def _df_to_parquet_codec(df: pd.DataFrame, path: Path) -> str:
    for codec in ("zstd", "snappy"):
        try:
            df.to_parquet(path, compression=codec, index=False)
            return codec
        except Exception:
            continue
    df.to_parquet(path, index=False)
    return "none"


@dataclass
class SheetWork:
    sheet_id: str
    display_name: str
    parquet_relative: str
    profile_relative: str
    preview_relative: str
    row_count: int
    col_count: int
    warnings: list[str] = field(default_factory=list)


@dataclass
class ParseOutcome:
    workbook_manifest: dict[str, Any]
    sheet_catalog: dict[str, Any]
    preview_md: str
    formulas: list[dict[str, Any]]
    named_ranges: list[dict[str, Any]]
    merged_cells: list[dict[str, Any]]
    sheets: list[SheetWork]
    duckdb_relative: str
    parse_status: str
    warnings: list[str]
    skipped_sheets: list[dict[str, Any]]
    sniff_info: dict[str, Any] | None = None


def _extract_formulas_xlsx(path: Path) -> list[dict[str, Any]]:
    import openpyxl

    formulas: list[dict[str, Any]] = []
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    wb_data = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for idx, ws in enumerate(wb.worksheets, start=1):
            ws_d = wb_data[ws.title]
            sid = sanitize_sheet_id(ws.title, idx)
            for row in ws.iter_rows():
                for cell in row:
                    if getattr(cell, "data_type", None) != "f":
                        continue
                    formula_text = cell.value
                    if formula_text is None:
                        continue
                    cached = ws_d[cell.coordinate].value
                    formulas.append(
                        {
                            "sheet": ws.title,
                            "sheet_id": sid,
                            "cell": cell.coordinate,
                            "formula": str(formula_text),
                            "cached_value": cached,
                            "value_state": "cached" if cached is not None else "unknown",
                        }
                    )
    finally:
        wb.close()
        wb_data.close()
    return formulas


def _merged_cells_xlsx(path: Path) -> list[dict[str, Any]]:
    import openpyxl

    out: list[dict[str, Any]] = []
    wb = openpyxl.load_workbook(path, read_only=False)
    try:
        for idx, ws in enumerate(wb.worksheets, start=1):
            sid = sanitize_sheet_id(ws.title, idx)
            for rng in ws.merged_cells.ranges:
                out.append({"sheet": ws.title, "sheet_id": sid, "range": str(rng)})
    finally:
        wb.close()
    return out


def _named_ranges_xlsx(path: Path) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        out: list[dict[str, Any]] = []
        try:
            for name, defn in wb.defined_names.items():
                out.append({"name": str(name), "definition": str(defn)})
        except Exception:
            for dn in getattr(wb.defined_names, "values", lambda: [])():
                out.append({"name": str(getattr(dn, "name", "")), "definition": str(dn)})
        return out
    finally:
        wb.close()


def parse_xlsx(path: Path, cfg: TabularAgentConfig) -> ParseOutcome:
    import openpyxl

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet_names = xl.sheet_names
    merged = _merged_cells_xlsx(path)
    formulas = _extract_formulas_xlsx(path)
    named_ranges = _named_ranges_xlsx(path)
    warnings: list[str] = []
    skipped: list[dict[str, Any]] = []
    sheets_out: list[SheetWork] = []
    catalog: list[dict[str, Any]] = []

    meta_wb = openpyxl.load_workbook(path, read_only=True)
    try:
        hidden = sum(1 for s in meta_wb.worksheets if s.sheet_state != "visible")
        n_total = len(meta_wb.worksheets)
    finally:
        meta_wb.close()

    for idx, sheet_name in enumerate(sheet_names, start=1):
        if idx > cfg.max_sheets_per_workbook:
            skipped.append({"sheet": sheet_name, "reason": "max_sheets_exceeded"})
            continue
        sheet_id = sanitize_sheet_id(sheet_name, idx)
        try:
            df = pd.read_excel(xl, sheet_name=sheet_name, header=None, dtype=object)
        except Exception as exc:
            skipped.append({"sheet": sheet_name, "reason": f"read_failed:{exc}"})
            continue
        df, _header_row, _ = _trim_dataframe(df)
        if df.empty or df.shape[1] == 0:
            skipped.append({"sheet": sheet_name, "reason": "empty_after_trim"})
            continue
        header_idx = 0
        data_start = 1
        colnames = [str(c) if c is not None and str(c).strip() else f"col_{i}" for i, c in enumerate(df.iloc[0])]
        df = df.iloc[1:].copy() if len(df) > 1 else pd.DataFrame(columns=colnames)
        df.columns = colnames[: df.shape[1]]
        sw_list: list[str] = []
        if df.shape[1] > cfg.wide_column_warning_threshold:
            sw_list.append(f"wide_table:{df.shape[1]}_columns")
        if any(m["sheet_id"] == sheet_id for m in merged):
            sw_list.append("merged_cells_present")

        pq_rel = f"sheets/{sheet_id}.parquet"
        prof_rel = f"sheets/{sheet_id}_profile.json"
        prev_rel = f"sheets/{sheet_id}_preview.md"

        profile = {
            "sheet_id": sheet_id,
            "header_detection": build_header_detection_v1(header_idx=header_idx, data_start=data_start),
            "shape": {"rows": int(df.shape[0]), "cols": int(df.shape[1])},
            "dtypes": {str(c): str(t) for c, t in df.dtypes.items()},
            "null_density": round(float(df.isna().mean().mean()) if df.size else 0.0, 5),
            "warnings": sw_list,
        }

        sheets_out.append(
            SheetWork(
                sheet_id=sheet_id,
                display_name=sheet_name,
                parquet_relative=pq_rel,
                profile_relative=prof_rel,
                preview_relative=prev_rel,
                row_count=int(df.shape[0]),
                col_count=int(df.shape[1]),
                warnings=sw_list,
            )
        )
        sheets_out[-1].__dict__["_df"] = df  # noqa: SLF001
        sheets_out[-1].__dict__["_profile"] = profile  # noqa: SLF001

        catalog.append(
            {
                "sheet_id": sheet_id,
                "display_name": sheet_name,
                "row_count": int(df.shape[0]),
                "column_count": int(df.shape[1]),
                "hidden": False,
                "inferred_types": profile["dtypes"],
                "header_summary": profile["header_detection"],
                "multi_table_warning": profile["header_detection"].get("multiple_table_warning", False),
                "merged_cell_warning": any(m["sheet_id"] == sheet_id for m in merged),
            }
        )

    xl.close()

    status = "parsed_with_warnings" if warnings or skipped else "parsed"
    if not catalog:
        status = "parse_failed" if not skipped else "partial_parse"
        warnings.append("no_sheets_parsed")

    manifest = {
        "filename": path.name,
        "format": "xlsx",
        "sheet_count": n_total,
        "visible_sheet_count": max(0, n_total - hidden),
        "hidden_sheet_count": hidden,
        "workbook_warnings": warnings,
    }
    preview_md = f"# {path.name}\n\n- Format: xlsx\n- Parsed sheets: {len(catalog)}\n- Skipped: {len(skipped)}\n"
    return ParseOutcome(
        workbook_manifest=manifest,
        sheet_catalog={"sheets": catalog, "skipped": skipped},
        preview_md=preview_md,
        formulas=formulas,
        named_ranges=named_ranges,
        merged_cells=merged,
        sheets=sheets_out,
        duckdb_relative="bundle.duckdb",
        parse_status=status,
        warnings=warnings,
        skipped_sheets=skipped,
    )


def parse_csv_tsv(path: Path, cfg: TabularAgentConfig, *, is_tsv: bool) -> ParseOutcome:
    encoding, sep = _sniff_csv_file(path)
    if is_tsv:
        sep = "\t"
    sniff_warn = False
    try:
        df = pd.read_csv(path, sep=sep, encoding=encoding, engine="python", on_bad_lines="warn", dtype=object)
    except TypeError:
        df = pd.read_csv(path, sep=sep, encoding=encoding, engine="python", dtype=object)
    except Exception:
        df = pd.read_csv(path, sep=sep, encoding="utf-8", engine="python", dtype=object)
        sniff_warn = True
    df, _, _ = _trim_dataframe(df)
    warnings = []
    if sniff_warn:
        warnings.append("csv_fallback_encoding")
    sheet_id = "sheet_1"
    sw = []
    if df.shape[1] > cfg.wide_column_warning_threshold:
        sw.append(f"wide_table:{df.shape[1]}_columns")

    profile = {
        "sheet_id": sheet_id,
        "header_detection": build_header_detection_v1(header_idx=0, data_start=1),
        "shape": {"rows": int(df.shape[0]), "cols": int(df.shape[1])},
        "dtypes": {str(c): str(t) for c, t in df.dtypes.items()},
        "null_density": round(float(df.isna().mean().mean()) if df.size else 0.0, 5),
        "warnings": sw,
    }
    sh = SheetWork(
        sheet_id=sheet_id,
        display_name="data",
        parquet_relative=f"sheets/{sheet_id}.parquet",
        profile_relative=f"sheets/{sheet_id}_profile.json",
        preview_relative=f"sheets/{sheet_id}_preview.md",
        row_count=int(df.shape[0]),
        col_count=int(df.shape[1]),
        warnings=sw,
    )
    sh.__dict__["_df"] = df  # noqa: SLF001
    sh.__dict__["_profile"] = profile  # noqa: SLF001

    sniff = {"encoding": encoding, "delimiter": sep, "is_tsv": is_tsv}
    manifest = {
        "filename": path.name,
        "format": "tsv" if is_tsv else "csv",
        "sheet_count": 1,
        "visible_sheet_count": 1,
        "hidden_sheet_count": 0,
        "workbook_warnings": warnings,
    }
    preview_md = f"# {path.name}\n\n- Format: {manifest['format']}\n- Rows: {df.shape[0]}\n"
    return ParseOutcome(
        workbook_manifest=manifest,
        sheet_catalog={"sheets": [catalog_entry_from_sheet(sh, profile)], "skipped": []},
        preview_md=preview_md,
        formulas=[],
        named_ranges=[],
        merged_cells=[],
        sheets=[sh],
        duckdb_relative="bundle.duckdb",
        parse_status="parsed_with_warnings" if warnings else "parsed",
        warnings=warnings,
        skipped_sheets=[],
        sniff_info=sniff,
    )


def catalog_entry_from_sheet(sh: SheetWork, profile: dict[str, Any]) -> dict[str, Any]:
    hd = profile.get("header_detection") if isinstance(profile.get("header_detection"), dict) else {}
    return {
        "sheet_id": sh.sheet_id,
        "display_name": sh.display_name,
        "row_count": sh.row_count,
        "column_count": sh.col_count,
        "hidden": False,
        "inferred_types": profile.get("dtypes", {}),
        "header_summary": profile.get("header_detection"),
        "multi_table_warning": bool(hd.get("multiple_table_warning", False)),
        "merged_cell_warning": False,
    }


def parse_xlsb(path: Path, cfg: TabularAgentConfig) -> ParseOutcome:
    xl = pd.ExcelFile(path, engine="pyxlsb")
    all_sheet_names = list(xl.sheet_names)
    merged: list[dict[str, Any]] = []
    formulas: list[dict[str, Any]] = []
    named_ranges: list[dict[str, Any]] = []
    warnings: list[str] = []
    skipped: list[dict[str, Any]] = []
    sheets_out: list[SheetWork] = []
    catalog: list[dict[str, Any]] = []

    for idx, sheet_name in enumerate(xl.sheet_names, start=1):
        if idx > cfg.max_sheets_per_workbook:
            skipped.append({"sheet": sheet_name, "reason": "max_sheets_exceeded"})
            continue
        sheet_id = sanitize_sheet_id(sheet_name, idx)
        try:
            df = pd.read_excel(xl, sheet_name=sheet_name, header=None, dtype=object, engine="pyxlsb")
        except Exception as exc:
            skipped.append({"sheet": sheet_name, "reason": f"read_failed:{exc}"})
            continue
        df, _, _ = _trim_dataframe(df)
        if df.empty:
            skipped.append({"sheet": sheet_name, "reason": "empty"})
            continue
        colnames = [str(c) if c is not None and str(c).strip() else f"col_{i}" for i, c in enumerate(df.iloc[0])]
        df = df.iloc[1:].copy() if len(df) > 1 else pd.DataFrame(columns=colnames)
        df.columns = colnames[: df.shape[1]]
        sw: list[str] = []
        if df.shape[1] > cfg.wide_column_warning_threshold:
            sw.append(f"wide_table:{df.shape[1]}_columns")
        profile = {
            "sheet_id": sheet_id,
            "header_detection": build_header_detection_v1(header_idx=0, data_start=1),
            "shape": {"rows": int(df.shape[0]), "cols": int(df.shape[1])},
            "dtypes": {str(c): str(t) for c, t in df.dtypes.items()},
            "null_density": round(float(df.isna().mean().mean()) if df.size else 0.0, 5),
            "warnings": sw,
        }
        sheets_out.append(
            SheetWork(
                sheet_id=sheet_id,
                display_name=sheet_name,
                parquet_relative=f"sheets/{sheet_id}.parquet",
                profile_relative=f"sheets/{sheet_id}_profile.json",
                preview_relative=f"sheets/{sheet_id}_preview.md",
                row_count=int(df.shape[0]),
                col_count=int(df.shape[1]),
                warnings=sw,
            )
        )
        sheets_out[-1].__dict__["_df"] = df  # noqa: SLF001
        sheets_out[-1].__dict__["_profile"] = profile  # noqa: SLF001
        catalog.append(catalog_entry_from_sheet(sheets_out[-1], profile))

    xl.close()
    manifest = {
        "filename": path.name,
        "format": "xlsb",
        "sheet_count": len(all_sheet_names),
        "visible_sheet_count": len(catalog),
        "hidden_sheet_count": 0,
        "workbook_warnings": warnings,
    }
    preview_md = f"# {path.name}\n\n- Format: xlsb\n- Parsed: {len(catalog)}\n"
    st = "parsed_with_warnings" if warnings or skipped else "parsed"
    if not catalog:
        st = "parse_failed"
        warnings.append("no_sheets_parsed")
    return ParseOutcome(
        workbook_manifest=manifest,
        sheet_catalog={"sheets": catalog, "skipped": skipped},
        preview_md=preview_md,
        formulas=formulas,
        named_ranges=named_ranges,
        merged_cells=merged,
        sheets=sheets_out,
        duckdb_relative="bundle.duckdb",
        parse_status=st,
        warnings=warnings,
        skipped_sheets=skipped,
    )


def persist_parse_outcome(
    *,
    bundle_root: Path,
    outcome: ParseOutcome,
    cfg: TabularAgentConfig,
) -> None:
    bundle_root.mkdir(parents=True, exist_ok=True)
    sheets_dir = bundle_root / "sheets"
    sheets_dir.mkdir(parents=True, exist_ok=True)
    (bundle_root / "codes").mkdir(exist_ok=True)
    (bundle_root / "executions").mkdir(exist_ok=True)
    (bundle_root / "assets" / "charts").mkdir(parents=True, exist_ok=True)
    (bundle_root / "assets" / "images").mkdir(parents=True, exist_ok=True)
    (bundle_root / "intermediate" / "normalized_input").mkdir(parents=True, exist_ok=True)

    for sh in outcome.sheets:
        df = getattr(sh, "_df", None)
        profile = getattr(sh, "_profile", {})
        if df is None:
            continue
        pq_path = bundle_root / sh.parquet_relative
        pq_path.parent.mkdir(parents=True, exist_ok=True)
        _df_to_parquet_codec(df, pq_path)
        (bundle_root / sh.profile_relative).write_text(
            json.dumps(profile, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        prev = df.iloc[: cfg.max_preview_rows, : cfg.max_preview_columns]
        try:
            md = prev.to_markdown(index=False)
        except Exception:
            md = prev.to_csv(index=False)
        (bundle_root / sh.preview_relative).write_text(md, encoding="utf-8")

    (bundle_root / "workbook_manifest.json").write_text(
        json.dumps(outcome.workbook_manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (bundle_root / "sheet_catalog.json").write_text(
        json.dumps(outcome.sheet_catalog, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (bundle_root / "preview.md").write_text(outcome.preview_md, encoding="utf-8")
    (bundle_root / "formulas.json").write_text(json.dumps(outcome.formulas, ensure_ascii=False, indent=2), encoding="utf-8")
    (bundle_root / "named_ranges.json").write_text(
        json.dumps(outcome.named_ranges, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (bundle_root / "merged_cells.json").write_text(
        json.dumps(outcome.merged_cells, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    if outcome.sniff_info:
        (bundle_root / "intermediate" / "normalized_input" / "csv_sniff.json").write_text(
            json.dumps(outcome.sniff_info, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    duck_path = bundle_root / outcome.duckdb_relative
    con = duckdb.connect(str(duck_path))
    try:
        for sh in outcome.sheets:
            pq = bundle_root / sh.parquet_relative
            if not pq.exists():
                continue
            view = f"s_{sh.sheet_id}"
            con.execute(f'DROP VIEW IF EXISTS "{view}"')
            con.execute(f'CREATE VIEW "{view}" AS SELECT * FROM read_parquet(?)', [str(pq.resolve())])
    finally:
        con.close()


def append_created_sheet_to_bundle(
    *,
    bundle_root: Path,
    sheet_id: str,
    display_name: str,
    df: pd.DataFrame,
    cfg: TabularAgentConfig,
) -> None:
    """
    Persist a new sheet's Parquet + profile + preview and refresh sheet_catalog.json,
    workbook_manifest.json, and preview.md. Caller must register the DuckDB view separately.
    """
    sheet_id = validate_safe_sheet_id(sheet_id)
    bundle_root = bundle_root.resolve()
    cat_path = bundle_root / "sheet_catalog.json"
    if not cat_path.is_file():
        raise FileNotFoundError(str(cat_path))
    catalog = json.loads(cat_path.read_text(encoding="utf-8"))
    sheets = catalog.setdefault("sheets", [])
    if any(isinstance(s, dict) and str(s.get("sheet_id")) == sheet_id for s in sheets):
        raise ValueError(f"duplicate sheet_id: {sheet_id}")

    wm_path = bundle_root / "workbook_manifest.json"
    manifest: dict[str, Any] = {}
    if wm_path.is_file():
        manifest = json.loads(wm_path.read_text(encoding="utf-8"))

    sw_list: list[str] = []
    if df.shape[1] > cfg.wide_column_warning_threshold:
        sw_list.append(f"wide_table:{df.shape[1]}_columns")

    hd = build_header_detection_v1(header_idx=0, data_start=1)
    hd = dict(hd)
    hd["notes"] = list(hd.get("notes") or []) + [
        "User-created sheet; column names were supplied explicitly (empty table until populated)."
    ]

    profile = {
        "sheet_id": sheet_id,
        "header_detection": hd,
        "shape": {"rows": int(df.shape[0]), "cols": int(df.shape[1])},
        "dtypes": {str(c): str(t) for c, t in df.dtypes.items()},
        "null_density": round(float(df.isna().mean().mean()) if df.size else 0.0, 5),
        "warnings": sw_list,
    }

    sheets_dir = bundle_root / "sheets"
    sheets_dir.mkdir(parents=True, exist_ok=True)
    pq_path = sheets_dir / f"{sheet_id}.parquet"
    _df_to_parquet_codec(df, pq_path)

    (bundle_root / "sheets" / f"{sheet_id}_profile.json").write_text(
        json.dumps(profile, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    prev = df.iloc[: cfg.max_preview_rows, : cfg.max_preview_columns]
    try:
        md = prev.to_markdown(index=False)
    except Exception:
        md = prev.to_csv(index=False)
    (bundle_root / "sheets" / f"{sheet_id}_preview.md").write_text(md, encoding="utf-8")

    entry: dict[str, Any] = {
        "sheet_id": sheet_id,
        "display_name": display_name,
        "row_count": int(df.shape[0]),
        "column_count": int(df.shape[1]),
        "hidden": False,
        "inferred_types": profile["dtypes"],
        "header_summary": hd,
        "multi_table_warning": False,
        "merged_cell_warning": False,
        "origin": "user_created",
    }
    sheets.append(entry)

    manifest["sheet_count"] = int(manifest.get("sheet_count") or 0) + 1
    manifest["visible_sheet_count"] = int(manifest.get("visible_sheet_count") or 0) + 1
    manifest.setdefault("workbook_warnings", [])
    wm_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    cat_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")

    pm = bundle_root / "preview.md"
    if pm.is_file():
        base = pm.read_text(encoding="utf-8").rstrip()
        extra = f"\n- Added sheet `{display_name}` (`{sheet_id}`), {int(df.shape[1])} column(s), {int(df.shape[0])} row(s).\n"
        pm.write_text(base + extra, encoding="utf-8")


def logical_bundle_paths(bundle_root: Path, artifacts_root: Path) -> dict[str, str]:
    resolved = bundle_root.resolve()
    relative_to_artifacts = resolved.relative_to(artifacts_root.resolve())
    base = (Path("runs") / "artifacts" / relative_to_artifacts).as_posix()

    def child(name: str) -> str:
        return f"{base}/{name}".replace("\\", "/")

    return {
        "bundle_root": base,
        "manifest": child("manifest.json"),
        "workbook_manifest": child("workbook_manifest.json"),
        "sheet_catalog": child("sheet_catalog.json"),
        "preview_md": child("preview.md"),
        "bundle_duckdb": child("bundle.duckdb"),
    }
